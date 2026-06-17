# -- coding: utf-8 --
# @Author: 胡H
# @File: app/services/ftp_sync_engine.py
# @Created: 2026/6/9 10:51
# @LastModified: 
# Copyright (c) 2026 by 胡H, All Rights Reserved.
# @desc: FTP 文件采集: 支持文件下载、MD5去重、结构化文件解析入库

import csv
import fnmatch
import hashlib
import json
import os
import time
from datetime import datetime
from ftplib import error_perm, error_reply
from pathlib import Path
from typing import Optional

import yaml
from sqlalchemy import Table, Column, Text, MetaData, String, JSON, inspect
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.core import logger, project_rootpath
from app.core.config import settings
from app.db.session import collected_engine as global_target_engine, SessionLocal
from app.models.taskLogModel import FtpFileRecord
from app.schemas.tsync import DBSyncReq
from app.services.task_control import get_task_status, TASK_PAUSED, TASK_CANCELLED
from app.exceptions import TaskPausedException, TaskCancelledException
from app.services.file_client_factory import FileClientFactory


class FtpSyncEngine:
    """
    FTP 文件采集
      - 单文件下载 + MD5 去重
      - 结构化文件解析入库(CSV / JSON / YAML)
      - 二进制/任意文件仅存储本地 + 记录元数据
    """

    DEFAULT_SAVE_DIR = Path(project_rootpath, settings.FTP_LOCAL_SAVE_DIR)
    DEFAULT_SAVE_DIR.mkdir(parents=True, exist_ok=True)

    def __init__(self, req: DBSyncReq, target_engine=global_target_engine):
        self.req = req
        self.target_engine = target_engine
        self.batch_size = getattr(settings, "BATCH_SIZE", 1000)
        self.file_client = None

        # 如果传了完整 URL, 自动解析覆盖各字段
        if getattr(self.req, "ftp_url", None):
            self._parse_url(self.req.ftp_url)

    def _parse_url(self, url: str):
        """
        解析多协议 URL, 覆盖 req 里的连接字段
        支持: ftp://  ftps://  sftp://  sdtp://
        """
        from urllib.parse import urlparse, unquote

        parsed = urlparse(url)
        scheme = parsed.scheme.lower()

        if scheme not in FileClientFactory.SUPPORTED_SCHEMES:
            raise ValueError(f"不支持的协议: {scheme}, 支持: {FileClientFactory.SUPPORTED_SCHEMES}")

        if parsed.hostname:
            self.req.host = parsed.hostname
        if parsed.port:
            self.req.port = parsed.port
        if parsed.username:
            self.req.username = unquote(parsed.username)
        if parsed.password:
            self.req.password = unquote(parsed.password)
        if parsed.path:
            # FTP/FTPS 虚拟目录需要去掉前导 / ; SFTP/SDTP 保留绝对路径
            self.req.ftp_path = parsed.path.lstrip("/") if scheme in ("ftp", "ftps") else parsed.path
        self.req.ftp_url = None  # 防止重复解析
        self.req.ftp_url_scheme = scheme  # 保存 scheme 供工厂路由

        logger.info(
            f"[{scheme}] URL 解析: host={self.req.host}, port={self.req.port}, user={self.req.username}, path={self.req.ftp_path}")

    #  FTP 连接

    def _connect_client(self):
        """ 通过工厂创建多协议客户端并连接 (FTP/FTPS/SFTP/SDTP) """
        scheme = getattr(self.req, 'ftp_url_scheme', None) or self.req.db_type or 'ftp'
        # 将 db_type=ftp 映射为默认 ftp（兼容纯表单配置）
        self.file_client = FileClientFactory.create(scheme)
        self.file_client.connect(
            host=self.req.host,
            port=self.req.port,
            username=self.req.username,
            password=self.req.password,
            passive=self.req.ftp_passive if getattr(self.req, "ftp_passive", None) is not None else True
        )

    #  MD5 计算

    def _compute_md5(self, filepath: str) -> str:
        """ 流式计算本地文件 MD5, 支持大文件 """
        h = hashlib.md5()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()

    #  文件下载

    def _build_local_path(self, remote_path: str) -> str:
        """
        根据远程路径生成本地存储路径
        规则: {save_dir}/{task_id}/{文件名}
        """
        save_dir = getattr(self.req, "local_save_dir", None) or self.DEFAULT_SAVE_DIR
        task_dir = os.path.join(save_dir, str(self.req.task_id))
        Path(task_dir).mkdir(parents=True, exist_ok=True)

        file_name = os.path.basename(remote_path)
        return os.path.join(task_dir, file_name)

    def _generate_row_id(self, row_dict: dict) -> str:
        """ 基于行内容的 MD5 哈希生成幂等主键, 重复执行不产生脏数据 """
        return hashlib.md5(json.dumps(row_dict, sort_keys=True, default=str).encode('utf-8')).hexdigest()

    #  JSONB 序列化清洗
    def _sanitize_for_jsonb(self, obj):
        """ 递归清洗对象, 确保所有值都能被 JSON 接受 """
        try:
            return json.loads(json.dumps(obj, default=str))
        except Exception:
            return str(obj)

    #  文件类型识别

    def _detect_file_type(self, file_name: str) -> str:
        """
        根据文件扩展名识别类型
        auto 模式下自动判断
        """
        file_type = getattr(self.req, "file_type", "auto") or "auto"
        if file_type != "auto":
            return file_type.lower()

        ext = Path(file_name).suffix.lower().lstrip(".")
        type_map = {
            "csv": "csv",
            "json": "json",
            "yaml": "yaml",
            "yml": "yaml",
            "xlsx": "xlsx",
            "xls": "xlsx",
            "xml": "xml",
        }
        return type_map.get(ext, "binary")

    #  结构化文件解析入库

    def _prepare_parse_target_table(self, table_name: str) -> Table:
        """
        为结构化文件内容准备目标表
        统一用两列结构: id(UUID PK) + raw_doc(JSON)
        如果表已存在则跳过建表
        """
        metadata = MetaData()
        table = Table(
            table_name,
            metadata,
            Column("id", String(32), primary_key=True),
            Column("raw_doc", JSON, nullable=False),
        )
        metadata.create_all(bind=self.target_engine, checkfirst=True)
        logger.info(f"解析目标表 [{table_name}] 已就绪")
        return table

    def _parse_csv(self, local_path: str) -> list:
        """
        解析 CSV 文件, 每行转为 dict
        自动处理 UTF-8 / GBK 编码
        """
        rows = []
        encodings = ["utf-8-sig", "gbk", "utf-8"]
        for encoding in encodings:
            try:
                with open(local_path, "r", encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        rows.append(dict(row))
                logger.info(f"CSV 解析成功, 编码: {encoding}, 共 {len(rows)} 行")
                return rows
            except UnicodeDecodeError:
                continue
        raise ValueError(f"CSV 文件编码识别失败, 尝试了: {encodings}")

    def _parse_json(self, local_path: str) -> list:
        """
        解析 JSON 文件
        支持顶层数组 [...] 和单个对象 {...}
        """
        with open(local_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            rows = data
        elif isinstance(data, dict):
            rows = [data]
        else:
            raise ValueError(f"不支持的 JSON 顶层结构: {type(data)}")

        logger.info(f"JSON 解析成功, 共 {len(rows)} 条")
        return rows

    def _parse_yaml(self, local_path: str) -> list:
        """
        解析 YAML 文件（支持多文档 YAML, 即 --- 分隔）
        """
        rows = []
        with open(local_path, "r", encoding="utf-8") as f:
            docs = list(yaml.safe_load_all(f))

        for doc in docs:
            if doc is None:
                continue
            if isinstance(doc, list):
                rows.extend(doc)
            elif isinstance(doc, dict):
                rows.append(doc)
            else:
                rows.append({"value": str(doc)})

        logger.info(f"YAML 解析成功, 共 {len(rows)} 条")
        return rows

    def _ingest_rows(self, rows: list, target_table_name: str) -> int:
        """
        将解析出的行列表批量写入目标表
        每行存为一个 JSON raw_doc
        """
        if not rows:
            return 0

        target_table = self._prepare_parse_target_table(target_table_name)
        total_inserted = 0

        with self.target_engine.begin() as conn:
            for i in range(0, len(rows), self.batch_size):
                # 探测暂停/取消信号
                self._check_task_status()

                batch = rows[i: i + self.batch_size]
                batch_data = [{"id": self._generate_row_id(row), "raw_doc": row} for row in batch]

                stmt = pg_insert(target_table).values(batch_data).on_conflict_do_nothing(index_elements=['id'])
                conn.execute(stmt)
                total_inserted += len(batch)

        return total_inserted

    def _parse_and_ingest(self, local_path: str, file_type: str) -> int:
        """
        解析结构化文件并写入目标表
        """
        target_table_name = getattr(self.req, "target_table", None)
        if not target_table_name:
            file_stem = Path(local_path).stem  # 文件名去掉扩展名
            target_table_name = f"ftp_{file_stem}"
            logger.info(f"未指定 target_table, 自动使用表名: {target_table_name}")

        logger.info(f"开始解析文件: {local_path}, 类型: {file_type}, 目标表: {target_table_name}")
        target_table = self._prepare_parse_target_table(target_table_name)
        total_inserted = 0

        # 流式处理 CSV (防 OOM
        if file_type == "csv":
            encodings = ["utf-8-sig", "gbk", "utf-8"]
            for encoding in encodings:
                try:
                    with open(local_path, "r", encoding=encoding) as f:
                        reader = csv.DictReader(f)
                        batch_data = []
                        with self.target_engine.begin() as conn:
                            for row in reader:
                                row_dict = dict(row)
                                batch_data.append(
                                    {"id": self._generate_row_id(row_dict),
                                     "raw_doc": self._sanitize_for_jsonb(row_dict)})
                                # 满一批次落盘并清空内存
                                if len(batch_data) >= self.batch_size:
                                    self._check_task_status()
                                    stmt = pg_insert(target_table).values(batch_data).on_conflict_do_nothing(
                                        index_elements=['id'])
                                    conn.execute(stmt)
                                    total_inserted += len(batch_data)
                                    batch_data.clear()
                            # 尾部处理
                            if batch_data:
                                stmt = pg_insert(target_table).values(batch_data).on_conflict_do_nothing(
                                    index_elements=['id'])
                                conn.execute(stmt)
                                total_inserted += len(batch_data)
                    logger.info(f"CSV 流式解析入库成功, 编码: {encoding}, 共 {total_inserted} 行")
                    return total_inserted
                except UnicodeDecodeError:
                    continue
            raise ValueError(f"CSV 文件编码识别失败, 尝试了: {encodings}")

        # 处理 JSON
        elif file_type == "json":
            with open(local_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            rows = data if isinstance(data, list) else [data]
            total_inserted = self._ingest_memory_rows(rows, target_table)
            logger.info(f"JSON 解析成功, 共 {total_inserted} 条")
            return total_inserted

        # 处理 YAML
        elif file_type == "yaml":
            rows = []
            with open(local_path, "r", encoding="utf-8") as f:
                docs = list(yaml.safe_load_all(f))
            for doc in docs:
                if doc is None: continue
                if isinstance(doc, list):
                    rows.extend(doc)
                elif isinstance(doc, dict):
                    rows.append(doc)
                else:
                    rows.append({"value": str(doc)})
            total_inserted = self._ingest_memory_rows(rows, target_table)
            logger.info(f"YAML 解析成功, 共 {total_inserted} 条")
            return total_inserted

        # 处理 Excel (.xlsx / .xls)
        elif file_type == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
            rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c or "") for c in row]
                        continue
                    rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
            wb.close()
            total_inserted = self._ingest_memory_rows(rows, target_table)
            logger.info(f"Excel 解析成功, {len(wb.sheetnames)} 个工作表, 共 {total_inserted} 条")
            return total_inserted

        # 处理 XML
        elif file_type == "xml":
            import xml.etree.ElementTree as ET
            tree = ET.parse(local_path)
            root = tree.getroot()

            def _xml_to_dict(elem):
                """ 递归将 XML 元素转为 dict """
                d = {"_tag": elem.tag}
                if elem.attrib:
                    d.update(elem.attrib)
                if elem.text and elem.text.strip():
                    d["_text"] = elem.text.strip()
                children = list(elem)
                if children:
                    for child in children:
                        child_dict = _xml_to_dict(child)
                        child_tag = child.tag
                        if child_tag in d:
                            if not isinstance(d[child_tag], list):
                                d[child_tag] = [d[child_tag]]
                            d[child_tag].append(child_dict)
                        else:
                            d[child_tag] = child_dict
                return d

            rows = [_xml_to_dict(root)]
            total_inserted = self._ingest_memory_rows(rows, target_table)
            logger.info(f"XML 解析成功, 共 {total_inserted} 条")
            return total_inserted

        else:
            logger.info(f"文件类型 [{file_type}] 不支持解析, 仅存储文件本身")
            return 0

    def _ingest_memory_rows(self, rows: list, target_table: Table) -> int:
        """ 将内存中的 list (JSON/YAML) 分批写入 """
        if not rows: return 0
        total = 0
        with self.target_engine.begin() as conn:
            for i in range(0, len(rows), self.batch_size):
                self._check_task_status()
                batch = [{"id": self._generate_row_id(row), "raw_doc": self._sanitize_for_jsonb(row)} for row in
                         rows[i: i + self.batch_size]]
                stmt = pg_insert(target_table).values(batch).on_conflict_do_nothing(index_elements=['id'])
                conn.execute(stmt)
                total += len(batch)
        return total

    #  数据库记录管理

    def _get_file_record(self, db, task_id: str, remote_path: str) -> Optional[FtpFileRecord]:
        """ 查询历史下载记录 """
        return db.query(FtpFileRecord).filter(
            FtpFileRecord.task_id == task_id,
            FtpFileRecord.remote_path == remote_path
        ).first()

    def _save_file_record(self, db, record_data: dict):
        """ 新增或更新文件记录 """
        existing = self._get_file_record(
            db, record_data["task_id"], record_data["remote_path"]
        )
        if existing:
            for k, v in record_data.items():
                setattr(existing, k, v)
            existing.update_time = datetime.now()
        else:
            record = FtpFileRecord(**record_data, downloaded_at=datetime.now())
            db.add(record)
        db.commit()

    #  状态探测

    def _check_task_status(self):
        task_id = str(self.req.task_id)
        status = get_task_status(task_id)
        if status == TASK_PAUSED:
            raise TaskPausedException("FTP 任务已暂停")
        if status == TASK_CANCELLED:
            raise TaskCancelledException("FTP 任务已取消")

    #  对外入口

    #  目录扫描

    def _is_sftp(self) -> bool:
        from app.services.file_client_factory import SftpClientAdapter
        return isinstance(self.file_client, SftpClientAdapter)

    def _list_dir(self, remote_dir: str) -> list:
        """获取指定目录下的文件/目录清单 (name, is_dir, mtime, size)"""
        items = []
        if self._is_sftp():
            for attr in self.file_client.sftp.listdir_attr(remote_dir):
                name = attr.filename
                if name in (".", ".."):
                    continue
                import stat as _stat
                is_dir = _stat.S_ISDIR(attr.st_mode)
                mtime = datetime.fromtimestamp(attr.st_mtime).strftime("%Y-%m-%d %H:%M:%S") if attr.st_mtime else None
                size = attr.st_size if not is_dir else 0
                items.append((name, is_dir, mtime, size))
        else:
            # FTP/FTPS: 优先用 MLSD，回退到 dir()
            ftp = self.file_client.ftp
            try:
                for name, facts in ftp.mlsd(remote_dir):
                    if name in (".", ".."):
                        continue
                    is_dir = facts.get("type") == "dir"
                    mtime = facts.get("modify")
                    try:
                        size = int(facts.get("size", 0))
                    except ValueError:
                        size = 0
                    items.append((name, is_dir, mtime, size))
            except (error_perm, error_reply):
                # 服务器不支持 MLSD，回退到 dir() 解析
                logger.info(f"MLSD 不可用，回退到传统 dir() 解析 [{remote_dir}]")
                lines = []
                ftp.dir(remote_dir, lines.append)
                for line in lines:
                    parts = line.split(None, 8)
                    if len(parts) < 9:
                        continue
                    name = parts[8]
                    if name in (".", ".."):
                        continue
                    info = parts[0]
                    is_dir = info.startswith("d")
                    size = int(parts[4]) if len(parts) > 4 else 0
                    # dir() 不返回 mtime，设为 None
                    items.append((name, is_dir, None, size))
        return items

    def _scan_directory(self, root_dir: str, is_recursive: bool) -> list:
        """
        DFS 递归扫描远程目录，返回扁平化的文件路径列表
        """
        file_paths = []

        def _scan(dir_path: str):
            self._check_task_status()
            try:
                items = self._list_dir(dir_path)
            except Exception as e:
                logger.warning(f"FTP 目录扫描失败 [{dir_path}]: {e}，跳过该目录")
                return

            dir_path = dir_path.rstrip("/")
            for name, is_dir, mtime, size in items:
                full_path = f"{dir_path}/{name}"
                if is_dir and is_recursive:
                    _scan(full_path)
                elif not is_dir:
                    file_paths.append({
                        "remote_path": full_path,
                        "file_name": name,
                        "remote_size": size,
                        "remote_mtime": mtime,
                    })

        _scan(root_dir.rstrip("/"))
        logger.info(f"FTP 目录扫描完成 [{root_dir}]，递归={is_recursive}, 发现 {len(file_paths)} 个文件")
        return file_paths

    #  单文件处理流程（单文件模式 / 批量模式复用）

    def _process_single_file(
            self, db, remote_path: str, file_name: str,
            remote_mtime: Optional[str], remote_size: Optional[int],
            target_table_name: str
    ) -> dict:
        """处理单个文件: 去重 → 下载 → 解析，返回该文件的统计"""
        task_id = str(self.req.task_id)
        local_path = self._build_local_path(remote_path)
        file_type = self._detect_file_type(file_name)

        existing = self._get_file_record(db, task_id, remote_path)

        # 快速去重: 远程 mtime + size 都对得上 → 跳过下载
        if existing and remote_mtime and remote_size:
            if (existing.remote_mtime == remote_mtime
                    and existing.remote_size == remote_size):
                logger.info(f"FTP 快速去重命中（mtime+size）, 跳过: {remote_path}")
                return {"status": "skipped_fast", "records": 0}

        # 下载 + MD5 去重
        file_size = self.file_client.download(remote_path, local_path, self._check_task_status)
        new_md5 = self._compute_md5(local_path)

        if existing and existing.md5 == new_md5:
            logger.info(f"FTP MD5 去重命中, 跳过: {remote_path}")
            return {"status": "skipped", "records": 0}

        self._save_file_record(db, {
            "task_id": task_id,
            "remote_path": remote_path,
            "local_path": local_path,
            "file_name": file_name,
            "file_size": file_size,
            "md5": new_md5,
            "remote_mtime": remote_mtime,
            "remote_size": remote_size,
            "file_type": file_type,
            "is_parsed": 0,
            "parsed_rows": 0,
        })

        parsed_rows = 0
        if getattr(self.req, "file_parse", False) and file_type != "binary":
            self._check_task_status()
            parsed_rows = self._parse_and_ingest(local_path, file_type)
            self._save_file_record(db, {
                "task_id": task_id,
                "remote_path": remote_path,
                "local_path": local_path,
                "file_name": file_name,
                "file_size": file_size,
                "md5": new_md5,
                "remote_mtime": remote_mtime,
                "remote_size": remote_size,
                "file_type": file_type,
                "is_parsed": 1,
                "parsed_rows": parsed_rows,
            })

        return {"status": "success", "records": parsed_rows, "size": file_size}

    #  对外入口

    def main(self) -> dict:
        ftp_dir = getattr(self.req, "ftp_dir", None)
        remote_path = getattr(self.req, "ftp_path", None) or self.req.db_name

        # 单文件模式: ftp_path 或 db_name 有值且不含通配特征
        is_batch = bool(ftp_dir)
        if not is_batch and not remote_path:
            raise ValueError("FTP 采集必须指定 ftp_path（单文件）或 ftp_dir（批量）")

        task_id = str(self.req.task_id)
        target_table_name = self.req.target_table or f"ftp_{Path(remote_path or ftp_dir or 'root').stem}"
        db = SessionLocal()
        start_time = time.time()

        try:
            self._connect_client()

            if is_batch:
                # ---- 批量模式 ----
                file_pattern = getattr(self.req, "file_pattern", "*") or "*"
                is_recursive = bool(getattr(self.req, "is_recursive", False))
                logger.info(
                    f"FTP 批量采集: dir={ftp_dir}, pattern={file_pattern}, recursive={is_recursive}"
                )

                candidates = self._scan_directory(ftp_dir, is_recursive)
                # 通配符过滤
                target_files = [
                    f for f in candidates
                    if fnmatch.fnmatch(f["file_name"], file_pattern)
                ]
                logger.info(
                    f"FTP 通配符过滤: {len(candidates)} 候选 → {len(target_files)} 匹配 (pattern={file_pattern})"
                )
            else:
                # ---- 单文件模式 ----
                target_files = [{
                    "remote_path": remote_path,
                    "file_name": os.path.basename(remote_path),
                    "remote_size": None,
                    "remote_mtime": None,
                }]

            if not target_files:
                logger.warning("未找到任何待采集的文件")
                return {
                    "status": "success", "tables_synced": 0, "total_records": 0,
                    "new_watermark": None, "table_details": []
                }

            table_details = []
            total_records = 0
            skipped = 0

            for f in target_files:
                self._check_task_status()
                try:
                    result = self._process_single_file(
                        db, f["remote_path"], f["file_name"],
                        f.get("remote_mtime"), f.get("remote_size"),
                        target_table_name,
                    )
                    if result["status"] in ("skipped", "skipped_fast"):
                        skipped += 1
                    else:
                        total_records += result["records"]
                        table_details.append({
                            "name": f["file_name"],
                            "target_name": target_table_name,
                            "records": result["records"],
                            "cost_seconds": None,
                            "high_watermark": None,
                        })
                except Exception as e:
                    logger.error(f"FTP 文件 [{f['remote_path']}] 处理失败: {e}")
                    continue

            elapsed = round(time.time() - start_time, 2)
            logger.info(
                f"FTP 采集完成，共 {len(target_files)} 个文件, "
                f"跳过 {skipped} 个, 写入 {total_records} 条, 耗时 {elapsed}s"
            )

            return {
                "status": "success",
                "tables_synced": 1 if total_records > 0 else 0,
                "total_records": total_records,
                "new_watermark": None,
                "table_details": table_details,
            }

        except (TaskPausedException, TaskCancelledException):
            raise

        except error_perm as e:
            logger.error(f"FTP 权限错误: {e}")
            raise RuntimeError(f"FTP 访问失败: {e}")

        except Exception as e:
            logger.error(f"FTP 采集异常: {e}")
            raise

        finally:
            if self.file_client:
                self.file_client.close()
            db.close()
