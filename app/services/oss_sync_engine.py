# -- coding: utf-8 --
# @Author: 胡H
# @File: app/services/oss_sync_engine.py
# @Created: 2026/6/17 9:31
# @LastModified: 
# Copyright (c) 2026 by 胡H, All Rights Reserved.
# @desc: OSS(S3兼容)采集-- 单文件/批量前缀下载 + MD5去重 + 结构化文件解析入库

import csv
import hashlib
import json
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import boto3
import yaml
from botocore.config import Config as BotoConfig
from botocore.exceptions import ClientError
from sqlalchemy import Table, Column, String, MetaData, JSON, Integer
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.core import logger, project_rootpath
from app.core.config import settings
from app.db.session import collected_engine as global_target_engine, SessionLocal
from app.models.taskLogModel import OssFileRecord
from app.schemas.tsync import DBSyncReq
from app.services.task_control import get_task_status, TASK_PAUSED, TASK_CANCELLED
from app.exceptions import TaskPausedException, TaskCancelledException


class OssSyncEngine:
    """
    OSS (S3兼容协议) 采集引擎
      - 单文件模式: 下载指定 object_key + MD5去重
      - 批量模式: 列举 prefix 下所有对象, 逐个下载处理
      - 结构化文件解析入库(CSV / JSON / YAML / Excel / XML) 
    """

    DEFAULT_SAVE_DIR = Path(project_rootpath, getattr(settings, "OSS_LOCAL_SAVE_DIR", "data/oss_files"))
    DEFAULT_SAVE_DIR.mkdir(parents=True, exist_ok=True)

    def __init__(self, req: DBSyncReq, target_engine=global_target_engine):
        self.req = req
        self.target_engine = target_engine
        self.batch_size = getattr(settings, "BATCH_SIZE", 1000)
        self.s3_client = None

    #  S3 客户端

    def _get_client(self):
        if self.s3_client:
            return self.s3_client

        addressing_style = self.req.oss_addressing_style or "virtual"

        # 匿名免密模式: 未传 AK/SK 时显式绕过签名, 避免 boto3 默认去环境变量找凭证
        if not self.req.oss_access_key or not self.req.oss_secret_key:
            from botocore import UNSIGNED
            logger.info("未检测到访问密钥, 自动切入 [OSS 匿名免密访问模式]")
            boto_config = BotoConfig(
                s3={"addressing_style": addressing_style},
                signature_version=UNSIGNED,
                retries={"max_attempts": 3, "mode": "standard"},
            )
            ak, sk = None, None
        else:
            boto_config = BotoConfig(
                s3={"addressing_style": addressing_style},
                signature_version="s3v4",
                retries={"max_attempts": 3, "mode": "standard"},
            )
            ak, sk = self.req.oss_access_key, self.req.oss_secret_key

        self.s3_client = boto3.client(
            "s3",
            endpoint_url=self.req.oss_endpoint,
            aws_access_key_id=ak,
            aws_secret_access_key=sk,
            region_name=self.req.oss_region or "us-east-1",
            use_ssl=bool(self.req.oss_use_ssl) if self.req.oss_use_ssl is not None else True,
            config=boto_config,
        )
        logger.info(f"OSS 客户端已创建: endpoint={self.req.oss_endpoint}, bucket={self.req.oss_bucket}")
        return self.s3_client

    #  状态探测

    def _check_task_status(self):
        task_id = str(self.req.task_id)
        status = get_task_status(task_id)
        if status == TASK_PAUSED:
            raise TaskPausedException("OSS 采集任务已暂停")
        if status == TASK_CANCELLED:
            raise TaskCancelledException("OSS 采集任务已取消")

    #  对象列举(批量模式) 

    def _list_objects(self, prefix: str) -> list:
        """
        列举 bucket 下指定前缀的所有对象 Key
        自动处理分页(超过max_keys时continuation_token翻页) 
        """
        client = self._get_client()
        bucket = self.req.oss_bucket
        max_keys = self.req.oss_max_keys or 1000

        object_keys = []
        continuation_token = None

        while True:
            self._check_task_status()

            params = {"Bucket": bucket, "Prefix": prefix, "MaxKeys": max_keys}
            if continuation_token:
                params["ContinuationToken"] = continuation_token

            resp = client.list_objects_v2(**params)
            contents = resp.get("Contents", [])

            for obj in contents:
                key = obj["Key"]
                # 排除"目录占位对象"(Key以/结尾, Size=0) 
                if key.endswith("/") and obj.get("Size", 0) == 0:
                    continue
                object_keys.append(key)

            if resp.get("IsTruncated"):
                continuation_token = resp.get("NextContinuationToken")
            else:
                break

        logger.info(f"前缀 [{prefix}] 下列举到 {len(object_keys)} 个对象")
        return object_keys

    #  文件下载

    def _build_local_path(self, object_key: str) -> str:
        """
        根据 object_key 生成本地存储路径
        {save_dir}/{task_id}/{object_key中的文件名}
        如果同一前缀下有多个同名文件(不同子目录) , 用key的hash加前缀区分, 避免互相覆盖
        """
        save_dir = getattr(self.req, "local_save_dir", None) or self.DEFAULT_SAVE_DIR
        task_dir = os.path.join(save_dir, str(self.req.task_id))
        Path(task_dir).mkdir(parents=True, exist_ok=True)

        file_name = os.path.basename(object_key)
        # 用 object_key 的短哈希作前缀, 防止同名文件互相覆盖
        key_hash = hashlib.md5(object_key.encode("utf-8")).hexdigest()[:8]
        return os.path.join(task_dir, f"{key_hash}_{file_name}")

    def _compute_md5(self, filepath: str) -> str:
        h = hashlib.md5()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()

    def _download_object(self, object_key: str, local_path: str) -> int:
        """
        下载单个对象到本地, 返回文件大小
        """
        client = self._get_client()
        bucket = self.req.oss_bucket

        try:
            client.download_file(bucket, object_key, local_path)
        except ClientError as e:
            error_code = e.response.get("Error", {}).get("Code", "")
            if error_code in ("404", "NoSuchKey"):
                raise RuntimeError(f"对象不存在: {object_key}")
            raise RuntimeError(f"OSS 下载失败 [{object_key}]: {e}")

        actual_size = os.path.getsize(local_path)
        if actual_size == 0:
            raise RuntimeError(f"对象下载后为空: {object_key}")

        logger.info(f"下载完成: {object_key} -> {local_path} ({actual_size} bytes)")
        return actual_size

    #  文件类型识别

    def _detect_file_type(self, file_name: str) -> str:
        file_type = getattr(self.req, "file_type", "auto") or "auto"
        if file_type != "auto":
            return file_type.lower()

        ext = Path(file_name).suffix.lower().lstrip(".")
        type_map = {
            "csv": "csv", "json": "json", "yaml": "yaml", "yml": "yaml",
            "xlsx": "xlsx", "xls": "xlsx", "xml": "xml",
        }
        return type_map.get(ext, "binary")

    #  JSON 序列化清洗 / 幂等ID

    def _sanitize_for_json(self, obj):
        try:
            return json.loads(json.dumps(obj, default=str))
        except Exception:
            return str(obj)

    def _generate_row_id(self, row_dict: dict) -> str:
        return hashlib.md5(json.dumps(row_dict, sort_keys=True, default=str).encode("utf-8")).hexdigest()

    #  结构化文件解析(复用FTP引擎的解析逻辑) 

    def _prepare_parse_target_table(self, table_name: str) -> Table:
        metadata = MetaData()
        table = Table(
            table_name,
            metadata,
            Column("id", String(32), primary_key=True),
            Column("source_key", String(500), nullable=True),  # 记录来源对象Key
            Column("raw_doc", JSON, nullable=False),
        )
        metadata.create_all(bind=self.target_engine, checkfirst=True)
        logger.info(f"解析目标表 [{table_name}] 已就绪")
        return table

    def _ingest_memory_rows(self, rows: list, target_table: Table, source_key: str) -> int:
        if not rows:
            return 0
        total = 0
        with self.target_engine.begin() as conn:
            for i in range(0, len(rows), self.batch_size):
                self._check_task_status()
                batch = [
                    {
                        "id": self._generate_row_id(row),
                        "source_key": source_key,
                        "raw_doc": self._sanitize_for_json(row)
                    }
                    for row in rows[i: i + self.batch_size]
                ]
                stmt = pg_insert(target_table).values(batch).on_conflict_do_nothing(index_elements=["id"])
                conn.execute(stmt)
                total += len(batch)
        return total

    def _parse_and_ingest(self, local_path: str, file_type: str, object_key: str, target_table_name: str) -> int:
        target_table = self._prepare_parse_target_table(target_table_name)

        if file_type == "csv":
            encodings = ["utf-8-sig", "gbk", "utf-8"]
            for encoding in encodings:
                try:
                    with open(local_path, "r", encoding=encoding) as f:
                        reader = csv.DictReader(f)
                        rows = [dict(row) for row in reader]
                    return self._ingest_memory_rows(rows, target_table, object_key)
                except UnicodeDecodeError:
                    continue
            raise ValueError(f"CSV 文件编码识别失败: {encodings}")

        elif file_type == "json":
            with open(local_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            rows = data if isinstance(data, list) else [data]
            return self._ingest_memory_rows(rows, target_table, object_key)

        elif file_type == "yaml":
            rows = []
            with open(local_path, "r", encoding="utf-8") as f:
                docs = list(yaml.safe_load_all(f))
            for doc in docs:
                if doc is None:
                    continue
                if isinstance(doc, list):
                    rows.extend(doc)
                elif isinstance(doc, dict):
                    rows.append(doc)
                else:
                    rows.append({"value": str(doc)})
            return self._ingest_memory_rows(rows, target_table, object_key)

        elif file_type == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
            rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c or "") for c in row]
                        continue
                    rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
            wb.close()
            return self._ingest_memory_rows(rows, target_table, object_key)

        elif file_type == "xml":
            import xml.etree.ElementTree as ET
            tree = ET.parse(local_path)
            root = tree.getroot()

            def _xml_to_dict(elem):
                d = {"_tag": elem.tag}
                if elem.attrib:
                    d.update(elem.attrib)
                if elem.text and elem.text.strip():
                    d["_text"] = elem.text.strip()
                for child in list(elem):
                    child_dict = _xml_to_dict(child)
                    tag = child.tag
                    if tag in d:
                        if not isinstance(d[tag], list):
                            d[tag] = [d[tag]]
                        d[tag].append(child_dict)
                    else:
                        d[tag] = child_dict
                return d

            rows = [_xml_to_dict(root)]
            return self._ingest_memory_rows(rows, target_table, object_key)

        else:
            logger.info(f"文件类型 [{file_type}] 不支持解析, 仅存储文件本身")
            return 0

    #  文件记录管理

    def _get_file_record(self, db, task_id: str, object_key: str) -> Optional[OssFileRecord]:
        return db.query(OssFileRecord).filter(
            OssFileRecord.task_id == task_id,
            OssFileRecord.object_key == object_key
        ).first()

    def _save_file_record(self, db, record_data: dict):
        existing = self._get_file_record(db, record_data["task_id"], record_data["object_key"])
        if existing:
            for k, v in record_data.items():
                setattr(existing, k, v)
            existing.update_time = datetime.now()
        else:
            record = OssFileRecord(**record_data, downloaded_at=datetime.now())
            db.add(record)
        db.commit()

    #  单对象处理流程(单文件模式 / 批量模式都调这个) 

    def _process_single_object(self, db, object_key: str, target_table_name: str) -> dict:
        """
        处理单个对象: 下载 → MD5去重 → 可选解析入库
        返回该对象的处理统计
        """
        file_name = os.path.basename(object_key)
        local_path = self._build_local_path(object_key)
        file_type = self._detect_file_type(file_name)
        task_id = str(self.req.task_id)

        existing_record = self._get_file_record(db, task_id, object_key)

        file_size = self._download_object(object_key, local_path)
        new_md5 = self._compute_md5(local_path)

        if existing_record and existing_record.md5 == new_md5:
            logger.info(f"对象未变更(MD5相同) , 跳过: {object_key}")
            return {"object_key": object_key, "status": "skipped", "records": 0}

        self._save_file_record(db, {
            "task_id": task_id,
            "object_key": object_key,
            "local_path": local_path,
            "file_name": file_name,
            "file_size": file_size,
            "md5": new_md5,
            "file_type": file_type,
            "is_parsed": 0,
            "parsed_rows": 0,
        })

        parsed_rows = 0
        if getattr(self.req, "file_parse", False) and file_type != "binary":
            self._check_task_status()
            parsed_rows = self._parse_and_ingest(local_path, file_type, object_key, target_table_name)
            self._save_file_record(db, {
                "task_id": task_id,
                "object_key": object_key,
                "local_path": local_path,
                "file_name": file_name,
                "file_size": file_size,
                "md5": new_md5,
                "file_type": file_type,
                "is_parsed": 1,
                "parsed_rows": parsed_rows,
            })

        return {"object_key": object_key, "status": "success", "records": parsed_rows, "size": file_size}

    def main(self) -> dict:
        if not self.req.oss_bucket:
            raise ValueError("OSS 采集必须指定 oss_bucket")

        object_key = getattr(self.req, "oss_object_key", None)
        prefix = getattr(self.req, "oss_prefix", None)

        if not object_key and not prefix:
            raise ValueError("OSS 采集必须指定 oss_object_key(单文件) 或 oss_prefix(批量) ")

        target_table_name = self.req.target_table or f"oss_{self.req.oss_bucket}"
        task_id = str(self.req.task_id)
        start_time = time.time()

        logger.info(f"启动 OSS 采集: bucket={self.req.oss_bucket}, endpoint={self.req.oss_endpoint}")

        db = SessionLocal()
        try:
            self._check_task_status()

            # 确定本次要处理的对象列表
            if object_key:
                # 单文件模式
                target_keys = [object_key]
            else:
                # 批量模式: 先列举
                target_keys = self._list_objects(prefix)

            if not target_keys:
                logger.warning("未找到任何待采集的对象")
                return {
                    "status": "success", "tables_synced": 0, "total_records": 0,
                    "new_watermark": None, "table_details": []
                }

            table_details = []
            total_records = 0
            skipped_count = 0

            for key in target_keys:
                self._check_task_status()
                try:
                    result = self._process_single_object(db, key, target_table_name)
                    if result["status"] == "skipped":
                        skipped_count += 1
                    else:
                        total_records += result["records"]
                        table_details.append({
                            "name": key,
                            "target_name": target_table_name,
                            "records": result["records"],
                            "cost_seconds": None,
                            "high_watermark": None
                        })
                except Exception as e:
                    logger.error(f"对象 [{key}] 处理失败: {e}")
                    # 单个对象失败不中断整个批次, 继续处理下一个
                    continue

            elapsed = round(time.time() - start_time, 2)
            logger.info(
                f"OSS 采集完成, 共 {len(target_keys)} 个对象, "
                f"跳过 {skipped_count} 个(未变更) , 写入 {total_records} 条记录, 耗时 {elapsed}s"
            )

            return {
                "status": "success",
                "tables_synced": 1 if total_records > 0 else 0,
                "total_records": total_records,
                "new_watermark": None,
                "table_details": table_details
            }

        except (TaskPausedException, TaskCancelledException):
            raise

        except Exception as e:
            logger.error(f"OSS 采集异常: {e}")
            raise

        finally:
            db.close()
